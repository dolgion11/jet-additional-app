# build_full_report_with_materiality.py
import calendar
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ========= SETTINGS =========
SRC_PATH = r"steppe road data.xlsx"      # <-- таны эх файл
TB_SHEET = "TB"
GL_SHEET = "GL"
MAT_SHEET = "Materiality"                # raw sheet байвал төгсгөлд оруулна
OUT_PATH = r"Full_JE_Report_with_Materiality.xlsx"

# Materiality inputs
CTT = 135_050_000     # Threshold (CT)
PM  = 1_620_600_000   # Performance Materiality (PM)
# ===========================

# ---------- helpers ----------
def pick(columns, options):
    norm = {str(c).strip().lower(): c for c in columns}
    for o in options:
        k = str(o).strip().lower()
        if k in norm: return norm[k]
    return None

THIN = Side(style="thin", color="A0A0A0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="F3F4F6")
BOLD = Font(bold=True)

def write_table(ws, df, start_row=1, start_col=1,
                money_cols=None, int_cols=None, percent_cols=None,
                freeze=True, show_grid=False, title_text=None, title_cell="B9"):
    money_cols   = set(money_cols   or [])
    int_cols     = set(int_cols     or [])
    percent_cols = set(percent_cols or [])
    # optional title (like “Testing:”)
    if title_text:
        t = ws[title_cell]
        t.value = title_text
        t.font  = Font(bold=True, color="CC0000")

    r = start_row
    # header
    for j, col in enumerate(df.columns, start=start_col):
        c = ws.cell(r, j, col); c.font=BOLD; c.fill=HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    r += 1
    # body
    for _, row in df.iterrows():
        for j, col in enumerate(df.columns, start=start_col):
            v = row[col]
            c = ws.cell(r, j, v); c.border = BORDER
            if col in money_cols:   c.number_format = '#,##0'
            if col in int_cols:     c.number_format = '0'
            if col in percent_cols: c.number_format = '0.0"%"'
        r += 1
    # autosize
    for j, col in enumerate(df.columns, start=start_col):
        w = min(max(12, max(len(str(col)), *(len(str(x)) for x in df[col])) + 2), 60)
        ws.column_dimensions[get_column_letter(j)].width = w
    if freeze:
        ws.freeze_panes = ws.cell(start_row+1, start_col)
    ws.sheet_view.showGridLines = show_grid
    return r-1

def ensure_abs(gl_df):
    g = gl_df.copy()
    if "ABS" in g.columns:
        g["ABS"] = pd.to_numeric(g["ABS"], errors="coerce").fillna(0)
    else:
        d = pick(g.columns, ["Debit","Дебет","Дебет дүн"])
        c = pick(g.columns, ["Credit","Кредит","Кредит дүн"])
        g["ABS"] = pd.to_numeric(g[d], errors="coerce").fillna(0).abs() + \
                   pd.to_numeric(g[c], errors="coerce").fillna(0).abs()
    return g

# ---------- load raw ----------
xls = pd.ExcelFile(SRC_PATH)
TB  = pd.read_excel(SRC_PATH, sheet_name=TB_SHEET) if TB_SHEET in xls.sheet_names else pd.DataFrame()
GL  = pd.read_excel(SRC_PATH, sheet_name=GL_SHEET) if GL_SHEET in xls.sheet_names else pd.DataFrame()
MAT_RAW = pd.read_excel(SRC_PATH, sheet_name=MAT_SHEET) if MAT_SHEET in xls.sheet_names else pd.DataFrame()

# ---------- 1) Reconciliation ----------
def build_reconciliation(TB, GL):
    if TB.empty or GL.empty: return pd.DataFrame()
    acc_tb = pick(TB.columns, ["Account No","Account number","Данс","Данс код"])
    c2023  = pick(TB.columns, ["2023", 2023])
    c2024  = pick(TB.columns, ["2024", 2024])

    tb = TB.copy()
    tb[acc_tb] = tb[acc_tb].astype(str).str.strip()
    tb[c2023]  = pd.to_numeric(tb[c2023], errors="coerce").fillna(0)
    tb[c2024]  = pd.to_numeric(tb[c2024], errors="coerce").fillna(0)
    tb_grp = (tb.groupby(acc_tb, as_index=False)
                .agg(Opening_TB=(c2023,"sum"), Ending_TB=(c2024,"sum")))

    first_digit = tb_grp[acc_tb].str[0]
    tb_grp["Movement_TB1"] = tb_grp["Ending_TB"].where(
        first_digit.isin(list("56789")),
        tb_grp["Ending_TB"] - tb_grp["Opening_TB"]
    )

    gl = GL.copy()
    acc_gl = pick(gl.columns, ["Account number","Данс"])
    txn    = pick(gl.columns, ["Transaction","Currency amount","Amount","Гүйлгээ","Дүн"])
    if txn is None:
        d = pick(gl.columns, ["Debit","Дебет","Дебет дүн"])
        c = pick(gl.columns, ["Credit","Кредит","Кредит дүн"])
        gl["__Transaction"] = pd.to_numeric(gl[d], errors="coerce").fillna(0) - \
                              pd.to_numeric(gl[c], errors="coerce").fillna(0)
        txn = "__Transaction"

    gl[acc_gl] = gl[acc_gl].astype(str).str.strip()
    gl[txn]    = pd.to_numeric(gl[txn], errors="coerce").fillna(0)
    gl_grp = (gl.groupby(acc_gl, as_index=False)
                .agg(Movement_GL=(txn,"sum"))
                .rename(columns={acc_gl: acc_tb}))

    df = (tb_grp.merge(gl_grp, how="left", on=acc_tb).fillna({"Movement_GL":0.0}))
    df["Difference Rounded"] = (df["Movement_TB1"] - df["Movement_GL"]).round(1)

    out = df.rename(columns={
        acc_tb: "Account No",
        "Opening_TB": "Opening Balance per TB",
        "Ending_TB": "Ending balance per TB (Total Correct)",
        "Movement_TB1": "Movement per TB 1",
        "Movement_GL": "Movement per GL"
    })
    total = {
        "Account No": "Total",
        "Opening Balance per TB": out["Opening Balance per TB"].sum(),
        "Ending balance per TB (Total Correct)": out["Ending balance per TB (Total Correct)"].sum(),
        "Movement per TB 1": out["Movement per TB 1"].sum(),
        "Movement per GL": out["Movement per GL"].sum(),
        "Difference Rounded": out["Difference Rounded"].sum()
    }
    return pd.concat([out, pd.DataFrame([total])], ignore_index=True)

# ---------- 2) Materiality (pretty, account_pivot_one_sheet style) ----------
def build_materiality(GL, CTT, PM):
    if GL.empty: return pd.DataFrame()
    g = ensure_abs(GL)

    intervals_info = [
        ("< 0", None, None),
        ("0 - 10% of Threshold",   "CTT", 0.10),
        ("10% - 20% of Threshold", "CTT", 0.20),
        ("20% - 30% of Threshold", "CTT", 0.30),
        ("30% - 40% of Threshold", "CTT", 0.40),
        ("40% - 50% of Threshold", "CTT", 0.50),
        ("50% - 60% of Threshold", "CTT", 0.60),
        ("60% - 70% of Threshold", "CTT", 0.70),
        ("70% - 80% of Threshold", "CTT", 0.80),
        ("80% - 90% of Threshold", "CTT", 0.90),
        ("90% of Threshold - Threshold", "CTT", 1.00),
        ("Threshold - 10% of PM",  "PM",  0.10),
        ("10% - 20% of PM",        "PM",  0.20),
        ("20% - 30% of PM",        "PM",  0.30),
        ("30% - 40% of PM",        "PM",  0.40),
        ("40% - 50% of PM",        "PM",  0.50),
        ("50% - 60% of PM",        "PM",  0.60),
        ("60% - 70% of PM",        "PM",  0.70),
        ("70% - 80% of PM",        "PM",  0.80),
        ("80% - 90% of PM",        "PM",  0.90),
        ("90% - 100% of PM",       "PM",  1.00),
        ("> 100% of PM",           "PM",  1.00)
    ]

    # interval numeric ranges (right edge values)
    interval_edges = []
    for _, ttype, pct in intervals_info:
        if ttype == "CTT":
            interval_edges.append(CTT * pct)
        elif ttype == "PM":
            interval_edges.append(PM * pct)
        else:
            interval_edges.append(None)

    # label for Amount Interval
    amount_labels = []
    for i, edge in enumerate(interval_edges):
        if i == 0:
            amount_labels.append("<0")
        else:
            prev = interval_edges[i-1] or 0
            curr = edge if edge is not None else prev
            amount_labels.append(f"{int(prev):,} - {int(curr):,}")

    # counts & totals
    counts, totals = [], []
    abs_series = pd.to_numeric(g["ABS"], errors="coerce")
    for i in range(len(intervals_info)):
        if i == 0:
            # exactly 0 (Excel COUNTIF "=" орчуулга)
            cnt = (abs_series == 0).sum()
            tot = abs_series[abs_series == 0].sum()
        elif i == len(intervals_info) - 1:
            lower = interval_edges[i-1]
            cnt = (abs_series > lower).sum()
            tot = abs_series[abs_series > lower].sum()
        else:
            lower = interval_edges[i-1] or 0
            upper = interval_edges[i]   or lower
            mask = ((abs_series > lower) & (abs_series < upper)) | (abs_series == upper)
            cnt = mask.sum()
            tot = abs_series[mask].sum()
        counts.append(int(cnt))
        totals.append(float(tot))

    total_count = sum(counts)
    total_amount = sum(totals)
    pct_count  = [ (c/total_count*100) if total_count else 0 for c in counts ]
    pct_amount = [ (t/total_amount*100) if total_amount else 0 for t in totals ]

    df = pd.DataFrame({
        "Materiality": [x[0] for x in intervals_info],
        "Amount Interval (in MNT)": amount_labels,
        "Interval range": interval_edges,
        "Number of Line Items Involved": counts,
        "Percentage": [round(x,1) for x in pct_count],
        "Total Amount (in MNT)": totals,
        "Amount Percentage": [round(x,1) for x in pct_amount]
    })
    # Total row
    df.loc[len(df)] = ["Total","", "", total_count, 100.0, total_amount, 100.0]
    return df

# ---------- 3) JE_by_Account (your pivot-one-sheet layout) ----------
def build_je_by_account_like_pivot(GL):
    g = ensure_abs(GL)
    col_accno = next(c for c in g.columns if c.lower() in
                     ["account number","данс","account number ".lower()])
    col_accnm = next(c for c in g.columns if c.lower() in
                     ["account name","дансны нэр","account name ".lower()])

    pivot = (g.groupby([col_accno, col_accnm], as_index=False)
               .agg(**{"Total Number of Entries": ("ABS","count"),
                       "Value of transactions": ("ABS","sum")})
               .rename(columns={col_accno:"Account Number", col_accnm:"Account Name"})
               .sort_values(["Account Number","Account Name"]).reset_index(drop=True))

    # totals
    total_entries = int(pivot["Total Number of Entries"].sum())
    total_value   = float(pivot["Value of transactions"].sum())

    # most / least
    max_entries = int(pivot["Total Number of Entries"].max())
    min_entries = int(pivot["Total Number of Entries"].min())
    most_list  = (pivot[pivot["Total Number of Entries"]==max_entries]
                  [["Account Name","Value of transactions"]]
                  .rename(columns={"Value of transactions":"Amount"})
                  .sort_values("Account Name").reset_index(drop=True))
    least_list = (pivot[pivot["Total Number of Entries"]==min_entries]
                  [["Account Name","Value of transactions"]]
                  .rename(columns={"Value of transactions":"Amount"})
                  .sort_values("Account Name").reset_index(drop=True))
    return pivot, total_entries, total_value, max_entries, min_entries, most_list, least_list

# ---------- 4) by month ----------
def build_by_month(GL):
    if GL.empty: return pd.DataFrame()
    g = ensure_abs(GL)
    dt = pick(g.columns, ["Date","Огноо"])
    g[dt] = pd.to_datetime(g[dt], errors="coerce"); g = g.dropna(subset=[dt])
    g["MonthNum"] = g[dt].dt.month; g["Month"] = g["MonthNum"].apply(lambda m: calendar.month_abbr[m])
    k = (g.groupby(["MonthNum","Month"], as_index=False)
          .agg(**{"Total Number of Line Items":("ABS","count"),
                  "Total Amount (in MNT)":("ABS","sum")})
          .sort_values("MonthNum"))
    df = k[["Month","Total Number of Line Items","Total Amount (in MNT)"]].reset_index(drop=True)
    df.loc[len(df)] = ["Total", df["Total Number of Line Items"].sum(), df["Total Amount (in MNT)"].sum()]
    return df

# ---------- 5) day group ----------
def build_day_group(GL):
    if GL.empty: return pd.DataFrame()
    g = ensure_abs(GL)
    day = pick(g.columns, ["Day","Өдөр"]); g["DayNum"] = pd.to_numeric(g[day], errors="coerce")
    def grp(x):
        if pd.isna(x): return None
        if 1 <= x <= 3:  return "<= 3 Days before M.E."
        if 4 <= x < 7:  return "4-7 Days before M.E."
        if 7 <= x < 14: return "8-14 Days before M.E."
        if x >= 14:     return "> 14 Days"
        return None
    g["Day Group"] = g["DayNum"].apply(grp)
    df = (g.groupby("Day Group", as_index=False)
            .agg(**{"Total Number of Line Items":("ABS","count"),
                    "Total Amount (in MNT)":("ABS","sum")}))
    total_items = df["Total Number of Line Items"].sum()
    df["%"] = (df["Total Number of Line Items"]/total_items*100).round(0).astype(int).astype(str)+"%"
    total = pd.DataFrame({"Day Group":["Total"],
                          "Total Number of Line Items":[total_items],
                          "%":["100%"],
                          "Total Amount (in MNT)":[df["Total Amount (in MNT)"].sum()]})
    return pd.concat([df,total], ignore_index=True)

# ---------- 6) by user ----------
def build_by_user(GL):
    if GL.empty: return pd.DataFrame()
    g = ensure_abs(GL)
    user = pick(g.columns, ["User","Бүртгэсэн хэрэглэгч"]); g[user] = g[user].fillna("(blank)")
    df = (g.groupby(user, as_index=False)
            .agg(**{"Total Number of Line Items":("ABS","count"),
                    "Total Amount (in MNT)":("ABS","sum")})
            .rename(columns={user:"User"})
            .reset_index(drop=True))
    tot = pd.DataFrame({"User":["Grand Total"],
                        "Total Number of Line Items":[df["Total Number of Line Items"].sum()],
                        "Total Amount (in MNT)":[df["Total Amount (in MNT)"].sum()]})
    return pd.concat([df,tot], ignore_index=True)

# ---------- 7) day of week ----------
def build_by_dow(GL):
    if GL.empty: return pd.DataFrame()
    g = GL.copy()
    dt = pick(g.columns, ["Үүсгэсэн огноо","Creation date"])
    g[dt] = pd.to_datetime(g[dt], errors="coerce"); g = g.dropna(subset=[dt])
    year = int(g[dt].dt.year.mode()[0]); days = 366 if calendar.isleap(year) else 365
    g["__wk"] = g[dt].dt.weekday + 1
    order = [(1,"Monday"),(2,"Tuesday"),(3,"Wednesday"),(4,"Thursday"),(5,"Friday"),(6,"Saturday"),(7,"Sunday")]
    rows=[]
    for k,name in order:
        cnt = int((g["__wk"]==k).sum()); rows.append([name,cnt,days,round(cnt/days,2)])
    df = pd.DataFrame(rows, columns=["Day","Total Number of Line Items","Days in the year","Average per day"])
    tot = pd.DataFrame({"Day":["Total"],
                        "Total Number of Line Items":[df["Total Number of Line Items"].sum()],
                        "Days in the year":[days],
                        "Average per day":[round(df["Total Number of Line Items"].sum()/days,0)]})
    return pd.concat([df,tot], ignore_index=True)

# ---------- 8) net to zero ----------
def build_net_to_zero(GL):
    if GL.empty: return pd.DataFrame()
    gl = GL.copy()
    acc = pick(gl.columns, ["Account number","Данс"])
    d   = pick(gl.columns, ["Debit","Дебет","Дебет дүн"])
    c   = pick(gl.columns, ["Credit","Кредит","Кредит дүн"])
    gl["__amt"] = pd.to_numeric(gl[d], errors="coerce").fillna(0) - pd.to_numeric(gl[c], errors="coerce").fillna(0)
    df = (gl.groupby(acc, as_index=False)["__amt"].sum()
            .rename(columns={acc:"Row Labels","__amt":"Sum of Transaction"}))
    return df[df["Sum of Transaction"]!=0].sort_values("Row Labels").reset_index(drop=True)

# ---------- build data ----------
recon            = build_reconciliation(TB, GL)
materiality_df   = build_materiality(GL, CTT, PM)     # 2-р sheet
pivot, total_entries, total_value, max_entries, min_entries, most_list, least_list = build_je_by_account_like_pivot(GL)
by_month         = build_by_month(GL)
by_day_group     = build_day_group(GL)
by_user          = build_by_user(GL)
by_dow           = build_by_dow(GL)
net_zero         = build_net_to_zero(GL)

# ---------- write excel ----------
wb = Workbook()

# 1) Reconcilation
ws = wb.active; ws.title = "Reconcilation"
write_table(ws, recon,
            money_cols={"Opening Balance per TB","Ending balance per TB (Total Correct)",
                        "Movement per TB 1","Movement per GL","Difference Rounded"})

# 2) Materiality (nice format)
ws = wb.create_sheet("Materiality")
last_r = write_table(
    ws, materiality_df,
    start_row=10, start_col=2,
    money_cols={"Total Amount (in MNT)"},
    int_cols={"Number of Line Items Involved"},
    percent_cols={"Percentage","Amount Percentage"},
    freeze=True, show_grid=False, title_text="Testing:"  # title style like your pivot
)

# 3) JE_by_Account (your one-sheet pivot layout)
ws = wb.create_sheet("JE_by_Account")
# Title
ws["B9"] = "Testing:"; ws["B9"].font = Font(bold=True, color="CC0000")
# Main pivot table
pivot_cols = ["Account Number","Account Name","Total Number of Entries","Value of transactions"]
end_r = write_table(ws, pivot[pivot_cols], start_row=10, start_col=2,
                    money_cols={"Value of transactions"},
                    int_cols={"Total Number of Entries"})
# Total row
tot_r = end_r + 1
ws.cell(tot_r, 4, "Total").font=BOLD; ws.cell(tot_r, 4).alignment=Alignment(horizontal="right")
ws.cell(tot_r, 5, total_entries).number_format='0'; ws.cell(tot_r,5).border=BORDER
ws.cell(tot_r, 6, total_value).number_format='#,##0'; ws.cell(tot_r,6).border=BORDER
# Comments
ws.cell(tot_r+2, 2, "Comments:").font=BOLD
# Summary header
hdr_r = tot_r + 4
for j, txt in enumerate(["Number of entries","Number of accounts","Account name","Amount"], start=3):
    c = ws.cell(hdr_r, j, txt); c.font=BOLD; c.fill=HDR_FILL
    c.alignment=Alignment(horizontal="center"); c.border=BORDER
# Most used
row = hdr_r + 2
ws.cell(row, 2, "Most used account").font=BOLD
ws.cell(row, 3, max_entries).number_format='0'; ws.cell(row,3).border=BORDER
ws.cell(row, 4, len(most_list)).number_format='0'; ws.cell(row,4).border=BORDER
for i, rec in most_list.iterrows():
    r = row + i
    ws.cell(r, 5, rec["Account Name"]).border=BORDER
    cc = ws.cell(r, 6, float(rec["Amount"])); cc.number_format='#,##0'; cc.border=BORDER
# Least used
row = hdr_r + 2 + max(1, len(most_list)) + 3
ws.cell(row, 2, "Least used accounts").font=BOLD
ws.cell(row, 3, min_entries).number_format='0'; ws.cell(row,3).border=BORDER
ws.cell(row, 4, len(least_list)).number_format='0'; ws.cell(row,4).border=BORDER
for i, rec in least_list.iterrows():
    r = row + i
    ws.cell(r, 5, rec["Account Name"]).border=BORDER
    cc = ws.cell(r, 6, float(rec["Amount"])); cc.number_format='#,##0'; cc.border=BORDER
ws.freeze_panes = "B11"; ws.sheet_view.showGridLines = False

# 4) JE_by_Month
ws = wb.create_sheet("JE_by_Month")
write_table(ws, by_month, money_cols={"Total Amount (in MNT)"},
            int_cols={"Total Number of Line Items"})

# 5) JE_by_Day_of_Month
ws = wb.create_sheet("JE_by_Day_of_Month")
write_table(ws, by_day_group, money_cols={"Total Amount (in MNT)"},
            int_cols={"Total Number of Line Items"})

# 6) JE Distribution by User
ws = wb.create_sheet("JE Distribution by User")
write_table(ws, by_user, money_cols={"Total Amount (in MNT)"},
            int_cols={"Total Number of Line Items"})

# 7) JE_by_Day_of_Week
ws = wb.create_sheet("JE_by_Day_of_Week")
write_table(ws, by_dow, int_cols={"Total Number of Line Items","Days in the year"})

# 8) Net_to_Zero_Test
ws = wb.create_sheet("Net_to_Zero_Test")
write_table(ws, net_zero, money_cols={"Sum of Transaction"})

# RAW sheets
ws = wb.create_sheet("TB");  write_table(ws, TB,  freeze=False, show_grid=True)
ws = wb.create_sheet("GL");  write_table(ws, GL,  freeze=False, show_grid=True)
ws = wb.create_sheet("materiality_raw"); write_table(ws, MAT_RAW, freeze=False, show_grid=True)

wb.save(OUT_PATH)
print(f"Done! Saved -> {Path(OUT_PATH).resolve()}")
